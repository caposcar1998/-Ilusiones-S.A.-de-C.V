from io import BytesIO
import json
import boto3
from botocore.exceptions import ClientError
from openpyxl import Workbook
import openpyxl
import logging
import os


def handler(event, context):
    nombre_archivo = event["body"]
    client = boto3.client('s3')

    res = client.get_object(
        Bucket = "m2crowdoscar",
        Key="ordenes/"+nombre_archivo
    )


    binary_data = res['Body'].read()
    wb = openpyxl.load_workbook(BytesIO(binary_data),data_only=True)
    sheet = wb.active
    column = 1
    for row in sheet.iter_rows():
        for cell in row:
            if(column == 1):
                column = column + 1
            else:
                column = column + 1
                sub_inventario = sheet.cell(column=1,row=column+1).value
                nombre = sheet.cell(column=2,row=column+1).value
                modelo = sheet.cell(column=3,row=column+1).value
                imei = sheet.cell(column=4,row=column+1).value
                folio = sheet.cell(column=5,row=column+1).value
                if(sub_inventario != None):
                    try:
                        print(sub_inventario)
                        update_almacen(sub_inventario,modelo,imei,folio)
                    except Exception as e:
                        print(e)
  
    return {
        'statusCode': 200,
        'headers': {
            'Access-Control-Allow-Headers': '*',
            'Access-Control-Allow-Origin': '*',
            'Access-Control-Allow-Methods': 'OPTIONS,POST,GET'
        },
        'body': json.dumps('Hello from your new Amplify Python lambda!')
    }

def update_almacen(sub_inventario,modelo,imei,folio):
    client = boto3.client('dynamodb')
    moS = str(modelo)
    FoS = str(folio)
    inS = str(imei)
    valor = {"M": 
    {
    "modelo":{"S":moS},
    "Imei":{"S":inS},
    "Folio":{"S":FoS},
    }} 

    response = client.update_item(
        TableName="Almacenes",
        Key={
            'sub_inventario':{"S":sub_inventario}
    },
        UpdateExpression="SET inventario = list_append(inventario, :i)",
        ExpressionAttributeValues={
            ":i": {"L":[valor]}
        }
    )