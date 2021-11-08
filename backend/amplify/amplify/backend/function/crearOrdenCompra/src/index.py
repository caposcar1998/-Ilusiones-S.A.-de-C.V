from io import BytesIO
import json
import boto3
from botocore.exceptions import ClientError
from openpyxl import Workbook
import openpyxl
import logging
import os
from datetime import date

def handler(event, context):
    nombre_archivo = event["body"]

    res = revisar_archivo(nombre_archivo)

    if (res == True):
        obtener_archivo_s3(nombre_archivo)
        return {
        'statusCode': 201,
        'headers': {
          'Access-Control-Allow-Headers': '*',
          'Access-Control-Allow-Origin': '*',
          'Access-Control-Allow-Methods': 'OPTIONS,POST,GET'
            },
        'body': json.dumps('Creado con exito')}
    else:
        return {
        'statusCode': 404,
        'headers': {
            'Access-Control-Allow-Headers': '*',
            'Access-Control-Allow-Origin': '*',
            'Access-Control-Allow-Methods': 'OPTIONS,POST,GET'
        },
        'body': "Archivo no valido"
        }

def obtener_archivo_s3(nombre_archivo):
    client = boto3.client('s3')

    res = client.get_object(
        Bucket = "m2crowdoscar",
        Key="ordenes/"+nombre_archivo
    )

    almacen_activo = True
    binary_data = res['Body'].read()
    wb = openpyxl.load_workbook(BytesIO(binary_data),data_only=True)
    ref = wb.active
    sheet1 = wb.active
    for row in sheet1.iter_rows():
        column = 1
        titulo = ""
        for cell in row:
            if(sheet1.cell(column=column,row=3) == "CERRADA POR CAMBIO DE LOCAL NO REALIZAR ENVIOS"):
                print("Uno no pasa")
            else:
                if(column == 1):
                    titulo = cell.value
                    workbook = Workbook()
                    sheet = workbook.active
                    sheet["A1"] = f"Orden de compra de:{cell.value}"
                    sub_inventario = cell.value
                    res = buscar_base_datos(cell.value)

                elif(column == 2):
                    sheet["A2"] = f"Cliente :{cell.value}"
                else:


                    if((cell.value != None)):
                        sheet.cell(column=5, row=sheet.max_row+1, value=str(ref.cell(row=1, column=column).value))
                        sheet.cell(column=6, row=sheet.max_row+1, value=cell.value)             
                column = column + 1   

    if(res == True):
        workbook.save(filename="/tmp/"+titulo+".xlsx")
        try:
            upload_file(("/tmp/"+titulo+".xlsx"),"m2crowdoscar","ordenes/ordenesFinales/"+titulo+".xlsx")
            update_ordenes(sub_inventario,f"https://m2crowdoscar.s3.us-west-2.amazonaws.com/ordenes/ordenesFinales/{titulo}.xlsx")
        except:
            pass
    else:
        logging.error("No encontrado en la base de datos")
    


def upload_file(file_name, bucket, object_name=None):

    s3 = boto3.client('s3')
    with open(file_name, "rb") as f:
        s3.upload_fileobj(f, bucket, object_name)

def revisar_archivo(nombre_archivo):
    client = boto3.client('s3')

    res = client.get_object(
        Bucket = "m2crowdoscar",
        Key="ordenes/"+nombre_archivo
    )

    binary_data = res['Body'].read()
    wb = openpyxl.load_workbook(BytesIO(binary_data),data_only=True)
    sheet = wb.active
    first_row = sheet[1]
    ans = False
    if (first_row[0].value == "Sub inventario" and first_row[1].value == "PDV"):
        ans = True

    return ans

def buscar_base_datos(subinventario):
    client = boto3.client('dynamodb')
    exist_file = False
    response = client.get_item(
        TableName="Almacenes",
        Key={
            "sub_inventario":{"S":subinventario} 
        }
    )
    try:
        res = response["Item"]
        exist_file = True
    except:
        exist_file = False        

    return exist_file

def update_ordenes(sub_inventario,archivo):
    client = boto3.client('dynamodb')
    today = date.today()
    s_today = str(today)
    valor = {"M": 
    {
    "dia":{"S":s_today},
    "archivo":{"S":archivo}
    }} 

    response = client.update_item(
        TableName="Almacenes",
        Key={
            'sub_inventario':{"S":sub_inventario}
    },
        UpdateExpression="SET ordenes = list_append(ordenes, :i)",
        ExpressionAttributeValues={
            ":i": {"L":[valor]}
        }
    )